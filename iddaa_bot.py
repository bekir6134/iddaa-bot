"""
İddaa Telegram Botu
- Excel'den veri okur (iddaa_analiz.xlsx)
- Maç analizi: takım bazlı ev/dep stats, İY, 2Y, gerçek oranlar
- Kupon önerisi, Alt/Üst kuponu, Handikap kuponu
- Her gün 08:00'de Excel otomatik güncellenir
"""
import os, re, logging
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (Application, CommandHandler, MessageHandler,
                           CallbackQueryHandler, ContextTypes, filters)
import openpyxl

TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN", "BURAYA_TOKEN")
SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH     = os.path.join(SCRIPT_DIR, "iddaa_analiz.xlsx")

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s - %(levelname)s - %(message)s")
log = logging.getLogger(__name__)

LIG_ISIMLERI = {
    "PL": "Premier League", "LL": "La Liga",  "BL": "Bundesliga",
    "SA": "Serie A",        "L1": "Ligue 1",  "SL": "Süper Lig",
}

ORANLAR_SAYFALAR    = ["PL - Oranlar","LL - Oranlar","BL - Oranlar",
                        "SA - Oranlar","L1 - Oranlar","SL - Oranlar"]
ISTATISTIK_SAYFALAR = ["PL - Istatistik","LL - Istatistik","BL - Istatistik",
                        "SA - Istatistik","L1 - Istatistik"]   # SL yok
PUAN_SAYFALAR       = ["PL - Puan Tablosu","LL - Puan Tablosu","BL - Puan Tablosu",
                        "SA - Puan Tablosu","L1 - Puan Tablosu"]
TAKIM_SAYFALAR      = ["PL - Takim","LL - Takim","BL - Takim",
                        "SA - Takim","L1 - Takim"]

# ── Takım adı normalize (FD ↔ Odds API eşleştirme) ───────────────────────────
def normalize_takim(isim):
    isim = re.sub(r'\b(FC|AFC|SC|CF|FK|SK|1\.)\b\.?', '', str(isim))
    return isim.strip().lower()

def takim_stat_bul(takim_stat_lig, isim):
    """Normalize edilmiş isimle takım istatistiğini bul (kısmi eşleştirme)."""
    norm = normalize_takim(isim)
    # Tam eşleşme
    if norm in takim_stat_lig:
        return takim_stat_lig[norm]
    # Kısmi eşleştirme
    for k, v in takim_stat_lig.items():
        if norm in k or k in norm:
            return v
    return {}

# ── Excel Okuyucu ─────────────────────────────────────────────────────────────
def excel_oku():
    if not os.path.exists(EXCEL_PATH):
        return None
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    veri = {"oranlar": [], "istatistik": {}, "puan": [],
            "takim_stat": {}, "guncelleme": ""}

    try:
        veri["guncelleme"] = str(wb["OZET"].cell(1,1).value or "")
    except Exception:
        pass

    def _f(v):
        try: return float(v)
        except (TypeError, ValueError): return 0.0

    # Oranlar — 22 sütun
    for sh in ORANLAR_SAYFALAR:
        if sh not in wb.sheetnames: continue
        lig = sh.split(" - ")[0]
        for row in wb[sh].iter_rows(min_row=3, values_only=True):
            if not row[0] or "Veri" in str(row[0]) or "Yaklaşan" in str(row[0]): continue
            try:
                veri["oranlar"].append({
                    "lig": lig, "tarih": str(row[0]), "saat": str(row[1] or ""),
                    "ev": str(row[2] or ""), "dep": str(row[3] or ""),
                    "o1": _f(row[4]),  "ox": _f(row[5]),  "o2": _f(row[6]),
                    "p1": _f(row[7]),  "px": _f(row[8]),  "p2": _f(row[9]),
                    "marj":    _f(row[10]), "favori": str(row[11] or ""),
                    "over25":  _f(row[12]), "under25":  _f(row[13]),
                    "kg_evet": _f(row[14]), "kg_hayir": _f(row[15]),
                    "iy1": _f(row[16]), "iyx": _f(row[17]), "iy2": _f(row[18]),
                    "hcp_ev":  _f(row[19]), "hcp_dep": _f(row[20]),
                    "hcp_val": str(row[21] or "-"),
                })
            except Exception:
                pass

    # İstatistik (lig geneli)
    for sh in ISTATISTIK_SAYFALAR:
        if sh not in wb.sheetnames: continue
        lig = sh.split(" - ")[0]
        stat = {}
        for row in wb[sh].iter_rows(min_row=3, values_only=True):
            if row[0] and row[1] and not str(row[0]).startswith("─"):
                stat[str(row[0])] = {"deger": row[1], "pct": row[2]}
        veri["istatistik"][lig] = stat

    # Puan tablosu
    for sh in PUAN_SAYFALAR:
        if sh not in wb.sheetnames: continue
        lig = sh.split(" - ")[0]
        for row in wb[sh].iter_rows(min_row=3, values_only=True):
            if row[0] and row[1]:
                veri["puan"].append({
                    "lig": lig, "sira": row[0], "takim": str(row[1]),
                    "oyun": row[2], "galibiyet": row[3], "beraberlik": row[4],
                    "maglubiyet": row[5], "puan": row[9], "form": str(row[10] or ""),
                })

    # Takım bazlı istatistik
    # Sütunlar: normalize_isim(0) | ev_mac(1) | ev_g%(2) | ev_att(3) | ev_yedi(4) |
    #           ev_ust15%(5) | ev_ust25%(6) | ev_kg%(7) |
    #           dep_mac(8) | dep_g%(9) | dep_att(10) | dep_yedi(11) |
    #           dep_ust15%(12) | dep_ust25%(13) | dep_kg%(14) |
    #           son5_form(15) | son5_att(16) | son5_yedi(17)
    for sh in TAKIM_SAYFALAR:
        if sh not in wb.sheetnames: continue
        lig = sh.split(" - ")[0]
        veri["takim_stat"][lig] = {}
        for row in wb[sh].iter_rows(min_row=3, values_only=True):
            if not row[0] or str(row[0]).startswith("─"): continue
            norm = str(row[0]).strip().lower()
            def pct_to_f(v):
                """'%80' veya 0.80 veya 80 → float"""
                try:
                    s = str(v).replace("%","").strip()
                    return float(s)
                except (ValueError, TypeError):
                    return 0.0
            veri["takim_stat"][lig][norm] = {
                "ev_mac":   _f(row[1]),
                "ev_g_pct": pct_to_f(row[2]),
                "ev_att":   _f(row[3]),
                "ev_yedi":  _f(row[4]),
                "ev_ust15": pct_to_f(row[5]),
                "ev_ust25": pct_to_f(row[6]),
                "ev_kg":    pct_to_f(row[7]),
                "dep_mac":  _f(row[8]),
                "dep_g_pct":pct_to_f(row[9]),
                "dep_att":  _f(row[10]),
                "dep_yedi": _f(row[11]),
                "dep_ust15":pct_to_f(row[12]),
                "dep_ust25":pct_to_f(row[13]),
                "dep_kg":   pct_to_f(row[14]),
                "son5_form":str(row[15] or "-"),
                "son5_att": _f(row[16]),
                "son5_yedi":_f(row[17]),
            }
    return veri

# ── Maç Ara ───────────────────────────────────────────────────────────────────
def mac_ara(veri, sorgu):
    sorgu_lower = sorgu.lower()
    return [m for m in veri["oranlar"]
            if sorgu_lower in m["ev"].lower() or sorgu_lower in m["dep"].lower()]

# ── Maç Analizi ──────────────────────────────────────────────────────────────
def mac_analiz_metni(m, veri):
    lig      = m["lig"]
    lig_tam  = LIG_ISIMLERI.get(lig, lig)
    stat     = veri["istatistik"].get(lig, {})
    takim_db = veri["takim_stat"].get(lig, {})

    # Lig geneli
    ev_gal_pct = (stat.get("Ev Sahibi Galibiyeti",{}).get("pct") or 0)
    ust25_pct  = (stat.get("2.5 Üst (3+ gol)",    {}).get("pct") or 0)
    kg_pct     = (stat.get("Karşılıklı Gol (KG)",  {}).get("pct") or 0)
    ort_gol    = (stat.get("Maç Başı Ort. Gol",     {}).get("deger") or 0)
    iy_ev_pct  = (stat.get("İY Ev Galibiyeti",       {}).get("pct") or 0)
    iy_ber_pct = (stat.get("İY Beraberlik",          {}).get("pct") or 0)
    iy_dep_pct = (stat.get("İY Deplasman Galibiyeti",{}).get("pct") or 0)
    y2_ev_pct  = (stat.get("2Y Ev Galibiyeti",       {}).get("pct") or 0)
    y2_ber_pct = (stat.get("2Y Beraberlik",          {}).get("pct") or 0)
    y2_dep_pct = (stat.get("2Y Deplasman Galibiyeti",{}).get("pct") or 0)

    # Takım bazlı
    ev_s  = takim_stat_bul(takim_db, m["ev"])
    dep_s = takim_stat_bul(takim_db, m["dep"])

    # 1X2 öneri
    max_p = max(m["p1"], m["px"], m["p2"])
    if max_p == m["p1"]:   oneri="1 (Ev)"; oneri_oran=m["o1"]
    elif max_p == m["px"]: oneri="X (Beraberlik)"; oneri_oran=m["ox"]
    else:                  oneri="2 (Dep)"; oneri_oran=m["o2"]

    # Alt/Üst öneri — gerçek oran önce
    if m["over25"] and m["under25"] and m["over25"] > 1:
        if m["over25"] < m["under25"]:
            au_oneri = f"OVER 2.5 @ `{m['over25']}`"
        else:
            au_oneri = f"UNDER 2.5 @ `{m['under25']}`"
    else:
        if ust25_pct >= 0.55:   au_oneri = "OVER 2.5 (lig bazlı)"
        elif ust25_pct <= 0.45: au_oneri = "UNDER 2.5 (lig bazlı)"
        else:                    au_oneri = "Belirsiz"

    # İlk Yarı öneri
    if m["iy1"] and m["iyx"] and m["iy2"]:
        min_iy = min(m["iy1"], m["iyx"], m["iy2"])
        if min_iy == m["iy1"]:   iy_oneri = f"İY-1 @ `{m['iy1']}`"
        elif min_iy == m["iyx"]: iy_oneri = f"İY-X @ `{m['iyx']}`"
        else:                     iy_oneri = f"İY-2 @ `{m['iy2']}`"
    else:
        iy_oneri = "-"

    # Güven
    if max_p >= 60:   guven = "🟢 Yüksek"
    elif max_p >= 48: guven = "🟡 Orta"
    else:             guven = "🔴 Düşük"

    metin = (
        f"⚽ *{m['ev']} vs {m['dep']}*\n"
        f"🏆 {lig_tam}  |  📅 {m['tarih']}  {m['saat']}\n"
        f"{'─'*34}\n"
        f"💹 *1X2 Oranları*\n"
        f"  1️⃣ Ev:  `{m['o1']:.2f}`  (%{m['p1']:.0f})\n"
        f"  ✖️ X:   `{m['ox']:.2f}`  (%{m['px']:.0f})\n"
        f"  2️⃣ Dep: `{m['o2']:.2f}`  (%{m['p2']:.0f})\n"
        f"  📊 Marj: %{m['marj']:.1f}\n"
    )

    # İlk Yarı
    if m["iy1"] and m["iyx"] and m["iy2"]:
        metin += (f"{'─'*34}\n"
                  f"⏱️ *İlk Yarı 1X2*\n"
                  f"  İY-1: `{m['iy1']:.2f}`  |  İY-X: `{m['iyx']:.2f}`  |  İY-2: `{m['iy2']:.2f}`\n")

    # Handicap
    if m["hcp_ev"] and m["hcp_dep"]:
        metin += (f"{'─'*34}\n"
                  f"⚖️ *Asian Handicap* ({m['hcp_val']})\n"
                  f"  Ev: `{m['hcp_ev']:.2f}`  |  Dep: `{m['hcp_dep']:.2f}`\n")

    # Alt/Üst + KG
    metin += f"{'─'*34}\n🎯 *Alt/Üst & KG*\n"
    if m["over25"] and m["under25"]:
        metin += f"  Over 2.5: `{m['over25']:.2f}`  |  Under 2.5: `{m['under25']:.2f}`\n"
    else:
        metin += f"  2.5 Üst (lig): %{ust25_pct*100:.0f}  |  Ort. gol: {ort_gol:.2f}\n"
    if m["kg_evet"] and m["kg_hayir"]:
        metin += f"  KG Evet: `{m['kg_evet']:.2f}`  |  KG Hayır: `{m['kg_hayir']:.2f}`\n"
    else:
        metin += f"  KG (lig): %{kg_pct*100:.0f}\n"

    # Takım bazlı stats
    metin += f"{'─'*34}\n"
    if ev_s:
        metin += (f"🏠 *{m['ev']}* (Evde, {ev_s['ev_mac']:.0f} maç)\n"
                  f"  Galibiyet: %{ev_s['ev_g_pct']:.0f}  |  Gol: {ev_s['ev_att']:.2f} att / {ev_s['ev_yedi']:.2f} yedi\n"
                  f"  1.5Üst: %{ev_s['ev_ust15']:.0f}  |  2.5Üst: %{ev_s['ev_ust25']:.0f}  |  KG: %{ev_s['ev_kg']:.0f}\n"
                  f"  Son 5: {ev_s['son5_form']}  (Att:{ev_s['son5_att']:.1f} Yedi:{ev_s['son5_yedi']:.1f})\n")
    else:
        metin += f"🏠 *{m['ev']}* — istatistik mevcut değil\n"

    if dep_s:
        metin += (f"✈️ *{m['dep']}* (Dışarıda, {dep_s['dep_mac']:.0f} maç)\n"
                  f"  Galibiyet: %{dep_s['dep_g_pct']:.0f}  |  Gol: {dep_s['dep_att']:.2f} att / {dep_s['dep_yedi']:.2f} yedi\n"
                  f"  1.5Üst: %{dep_s['dep_ust15']:.0f}  |  2.5Üst: %{dep_s['dep_ust25']:.0f}  |  KG: %{dep_s['dep_kg']:.0f}\n"
                  f"  Son 5: {dep_s['son5_form']}  (Att:{dep_s['son5_att']:.1f} Yedi:{dep_s['son5_yedi']:.1f})\n")
    else:
        metin += f"✈️ *{m['dep']}* — istatistik mevcut değil\n"

    # Lig özeti
    if stat:
        metin += (f"{'─'*34}\n"
                  f"📊 *{lig_tam} Sezon Ortalaması*\n"
                  f"  İY: %{iy_ev_pct*100:.0f} Ev / %{iy_ber_pct*100:.0f} Ber / %{iy_dep_pct*100:.0f} Dep\n"
                  f"  2Y: %{y2_ev_pct*100:.0f} Ev / %{y2_ber_pct*100:.0f} Ber / %{y2_dep_pct*100:.0f} Dep\n"
                  f"  KG: %{kg_pct*100:.0f}  |  2.5Üst: %{ust25_pct*100:.0f}  |  Ort gol: {ort_gol:.2f}\n")

    metin += (f"{'─'*34}\n"
              f"🎯 *Öneri*\n"
              f"  1X2: *{oneri}* @ `{oneri_oran:.2f}`  {guven} (%{max_p:.0f})\n"
              f"  Alt/Üst: *{au_oneri}*\n"
              f"  İlk Yarı: *{iy_oneri}*\n")
    return metin

# ── Kupon Önerisi ─────────────────────────────────────────────────────────────
def kupon_oneri(veri, adet=3):
    import functools
    adaylar = []
    for m in veri["oranlar"]:
        max_p = max(m["p1"], m["px"], m["p2"])
        if max_p == m["p1"]:   sec="1"; oran=m["o1"]
        elif max_p == m["px"]: sec="X"; oran=m["ox"]
        else:                  sec="2"; oran=m["o2"]
        if max_p >= 48 and oran >= 1.15:  # eşik düşürüldü
            adaylar.append({**m, "sec": sec, "oran": oran, "ihtimal": max_p})

    adaylar.sort(key=lambda x: (x["ihtimal"] - x["marj"]), reverse=True)
    secilen = adaylar[:adet]

    if not secilen:
        return (f"⚠️ Yeterli güvenilir maç bulunamadı.\n\n"
                f"Şu an {len(veri['oranlar'])} yaklaşan maç var.\n"
                f"En yüksek ihtimal: %{max((max(m['p1'],m['px'],m['p2']) for m in veri['oranlar']), default=0):.0f}")

    toplam_oran = round(functools.reduce(lambda a,b: a*b, [s["oran"] for s in secilen]), 2)
    birlesik    = round(functools.reduce(lambda a,b: a*b, [s["ihtimal"]/100 for s in secilen])*100, 1)

    metin = f"🎫 *{adet} Maçlık Kupon Önerisi*\n{'─'*32}\n"
    for i, s in enumerate(secilen, 1):
        lig_tam = LIG_ISIMLERI.get(s["lig"], s["lig"])
        metin += (f"{i}. *{s['ev']} vs {s['dep']}*\n"
                  f"   📅 {s['tarih']}  |  🏆 {lig_tam}\n"
                  f"   Seçim: *{s['sec']}* @ `{s['oran']:.2f}`  (%{s['ihtimal']:.0f})\n\n")

    metin += (f"{'─'*32}\n"
              f"💹 Toplam Oran: *{toplam_oran}*\n"
              f"🎯 Birleşik İhtimal: *%{birlesik}*\n\n"
              f"💵 100 TL → *{round(100*toplam_oran)} TL*\n"
              f"💵 200 TL → *{round(200*toplam_oran)} TL*\n")
    return metin

# ── Alt/Üst Kuponu ────────────────────────────────────────────────────────────
def altust_kupon(veri, adet=3):
    adaylar = []
    for m in veri["oranlar"]:
        if m["over25"] and m["under25"] and m["over25"] > 1:
            if m["over25"] < m["under25"]:
                adaylar.append({**m, "au":"ÜST", "au_oran":m["over25"],
                                "au_pct": round(1/m["over25"]*100,1)})
            else:
                adaylar.append({**m, "au":"ALT", "au_oran":m["under25"],
                                "au_pct": round(1/m["under25"]*100,1)})
        else:
            lig = m["lig"]
            stat = veri["istatistik"].get(lig, {})
            ust_pct = (stat.get("2.5 Üst (3+ gol)",{}).get("pct") or 0) * 100
            if ust_pct >= 55:
                adaylar.append({**m,"au":"ÜST","au_oran":None,"au_pct":ust_pct})
            elif ust_pct <= 42:
                adaylar.append({**m,"au":"ALT","au_oran":None,"au_pct":100-ust_pct})

    adaylar.sort(key=lambda x: x["au_pct"], reverse=True)
    secilen = adaylar[:adet]

    if not secilen:
        return "⚠️ Yeterli alt/üst adayı bulunamadı."

    metin = f"📊 *{adet} Maçlık Alt/Üst Kuponu*\n{'─'*32}\n"
    for i, s in enumerate(secilen, 1):
        emoji   = "🔼" if s["au"] == "ÜST" else "🔽"
        lig_tam = LIG_ISIMLERI.get(s["lig"], s["lig"])
        oran_str = f"@ `{s['au_oran']:.2f}`" if s["au_oran"] else "(lig bazlı)"
        metin += (f"{i}. *{s['ev']} vs {s['dep']}*\n"
                  f"   📅 {s['tarih']}  |  🏆 {lig_tam}\n"
                  f"   {emoji} 2.5 *{s['au']}* {oran_str}  (%{s['au_pct']:.0f})\n\n")

    metin += f"{'─'*32}\n⚠️ Oranları bahisçiden kontrol et."
    return metin

# ── Handikap Kuponu ───────────────────────────────────────────────────────────
def handicap_kupon(veri, adet=3):
    adaylar = []
    for m in veri["oranlar"]:
        if not m["hcp_ev"] or not m["hcp_dep"]: continue
        if m["hcp_ev"] < m["hcp_dep"]:
            sec="Ev Hcp"; oran=m["hcp_ev"]
        else:
            sec="Dep Hcp"; oran=m["hcp_dep"]
        pct = round(1/oran*100, 1) if oran > 1 else 0
        if pct >= 50 and oran >= 1.65:
            adaylar.append({**m, "sec":sec, "oran":oran, "pct":pct})

    adaylar.sort(key=lambda x: (x["pct"]-x["marj"]), reverse=True)
    secilen = adaylar[:adet]

    if not secilen:
        return "⚠️ Yeterli handikap adayı bulunamadı."

    import functools
    toplam_oran = round(functools.reduce(lambda a,b: a*b, [s["oran"] for s in secilen]), 2)

    metin = f"🔀 *{adet} Maçlık Handikap Kuponu*\n{'─'*32}\n"
    for i, s in enumerate(secilen, 1):
        lig_tam = LIG_ISIMLERI.get(s["lig"], s["lig"])
        metin += (f"{i}. *{s['ev']} vs {s['dep']}*\n"
                  f"   📅 {s['tarih']}  |  🏆 {lig_tam}\n"
                  f"   Hcp ({s['hcp_val']}): *{s['sec']}* @ `{s['oran']:.2f}`  (%{s['pct']:.0f})\n\n")

    metin += (f"{'─'*32}\n"
              f"💹 Toplam Oran: *{toplam_oran}*\n"
              f"💵 100 TL → *{round(100*toplam_oran)} TL*\n"
              f"💵 200 TL → *{round(200*toplam_oran)} TL*\n"
              f"⚠️ Handikap bahisleri yüksek risk içerir.")
    return metin

# ── Ana Menü ─────────────────────────────────────────────────────────────────
def ana_menu_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⚽ Maç Analizi",   callback_data="mac_analizi"),
         InlineKeyboardButton("🎫 Kupon Öner",    callback_data="kupon_3")],
        [InlineKeyboardButton("📊 Alt/Üst Kupon", callback_data="altust"),
         InlineKeyboardButton("🔀 Handikap",      callback_data="handicap")],
        [InlineKeyboardButton("🏆 Puan Tablosu",  callback_data="puan")],
    ])

# ── Handler'lar ───────────────────────────────────────────────────────────────
async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    veri = excel_oku()
    gun  = veri["guncelleme"] if veri else "Excel bulunamadı"
    await update.message.reply_text(
        f"👋 *İddaa Analiz Botuna Hoş Geldin!*\n\n"
        f"📅 Son güncelleme: {gun}\n\n"
        f"Ne yapmak istiyorsun?",
        reply_markup=ana_menu_kb(), parse_mode="Markdown")

async def button_handler(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    veri = excel_oku()

    if not veri:
        await q.edit_message_text("❌ Excel bulunamadı. Önce `iddaa_analiz.py` çalıştır.")
        return

    kb_ana = [[InlineKeyboardButton("🏠 Ana Menü", callback_data="ana_menu")]]

    if q.data == "kupon_3":
        metin = kupon_oneri(veri, 3)
        kb = [[InlineKeyboardButton("5 Maçlık", callback_data="kupon_5"),
               InlineKeyboardButton("🏠 Ana Menü", callback_data="ana_menu")]]
        await q.edit_message_text(metin, reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown")

    elif q.data == "kupon_5":
        metin = kupon_oneri(veri, 5)
        kb = [[InlineKeyboardButton("3 Maçlık", callback_data="kupon_3"),
               InlineKeyboardButton("🏠 Ana Menü", callback_data="ana_menu")]]
        await q.edit_message_text(metin, reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown")

    elif q.data == "altust":
        await q.edit_message_text(altust_kupon(veri, 3),
                                   reply_markup=InlineKeyboardMarkup(kb_ana), parse_mode="Markdown")

    elif q.data == "handicap":
        await q.edit_message_text(handicap_kupon(veri, 3),
                                   reply_markup=InlineKeyboardMarkup(kb_ana), parse_mode="Markdown")

    elif q.data == "puan":
        metin = "🏆 *Puan Tabloları (İlk 5)*\n\n"
        ligden = {}
        for t in veri["puan"]:
            ligden.setdefault(t["lig"], []).append(t)
        for kisa, tam in LIG_ISIMLERI.items():
            takimlar = ligden.get(kisa, [])
            if not takimlar: continue
            metin += f"*{tam}*\n"
            for t in takimlar[:5]:
                metin += f"  {t['sira']}. {t['takim'][:20]:<20} {t['puan']} puan  {t['form']}\n"
            metin += "\n"
        await q.edit_message_text(metin, reply_markup=InlineKeyboardMarkup(kb_ana), parse_mode="Markdown")

    elif q.data == "mac_analizi":
        await q.edit_message_text(
            "⚽ *Maç Analizi*\n\nTakım adını yaz:\n`Galatasaray`, `Arsenal`, `Bayern`, `Inter`, `PSG`",
            reply_markup=InlineKeyboardMarkup(kb_ana), parse_mode="Markdown")

    elif q.data == "ana_menu":
        gun = veri["guncelleme"]
        await q.edit_message_text(
            f"👋 *Ana Menü*\n📅 Son güncelleme: {gun}\n\nNe yapmak istiyorsun?",
            reply_markup=ana_menu_kb(), parse_mode="Markdown")

    elif q.data.startswith("mac_"):
        idx  = int(q.data.split("_")[1])
        veri2 = excel_oku()
        metin = mac_analiz_metni(veri2["oranlar"][idx], veri2)
        await q.edit_message_text(metin, reply_markup=InlineKeyboardMarkup(kb_ana), parse_mode="Markdown")

async def mesaj_handler(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sorgu = update.message.text.strip()
    if len(sorgu) < 2: return

    veri = excel_oku()
    if not veri:
        await update.message.reply_text("❌ Excel bulunamadı.")
        return

    bulunanlar = mac_ara(veri, sorgu)

    if not bulunanlar:
        await update.message.reply_text(
            f"🔍 `{sorgu}` için yaklaşan maç bulunamadı.\n\n"
            f"Şu an *{len(veri['oranlar'])}* yaklaşan maç mevcut.\n"
            f"Takım adı yaz: `Galatasaray`, `Arsenal`, `Bayern`, `Inter`, `PSG`",
            parse_mode="Markdown")
        return

    kb_ana = [[InlineKeyboardButton("🏠 Ana Menü", callback_data="ana_menu")]]
    if len(bulunanlar) == 1:
        metin = mac_analiz_metni(bulunanlar[0], veri)
        await update.message.reply_text(metin, reply_markup=InlineKeyboardMarkup(kb_ana),
                                         parse_mode="Markdown")
    else:
        metin = f"🔍 *'{sorgu}' için {len(bulunanlar)} maç:*\n\n"
        kb_rows = []
        for i, m in enumerate(bulunanlar[:8]):
            idx = veri["oranlar"].index(m)
            lig_tam = LIG_ISIMLERI.get(m["lig"], m["lig"])
            metin += f"{i+1}. {m['ev']} vs {m['dep']} ({m['tarih']}) — {lig_tam}\n"
            kb_rows.append([InlineKeyboardButton(
                f"{m['ev']} vs {m['dep']}", callback_data=f"mac_{idx}")])
        kb_rows.append([InlineKeyboardButton("🏠 Ana Menü", callback_data="ana_menu")])
        await update.message.reply_text(metin, reply_markup=InlineKeyboardMarkup(kb_rows),
                                         parse_mode="Markdown")

# ── Otomatik Güncelleme ───────────────────────────────────────────────────────
async def otomatik_guncelle(ctx: ContextTypes.DEFAULT_TYPE):
    log.info("Otomatik güncelleme başlatılıyor...")
    try:
        import subprocess
        subprocess.run(["python", os.path.join(SCRIPT_DIR, "iddaa_analiz.py")], timeout=300)
        log.info("Excel güncellendi!")
    except Exception as e:
        log.error(f"Otomatik güncelleme hatası: {e}")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("="*50)
    print("  İddaa Telegram Botu Başlatılıyor...")
    print("="*50)

    if not os.path.exists(EXCEL_PATH):
        print("⚠️  Excel bulunamadı, oluşturuluyor...")
        try:
            import subprocess
            subprocess.run(["python", os.path.join(SCRIPT_DIR, "iddaa_analiz.py")], timeout=300)
            print("✅ Excel oluşturuldu!")
        except Exception as e:
            print(f"❌ Excel oluşturulamadı: {e}")

    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("menu",  start))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, mesaj_handler))
    # Türkiye saati 07:00 = UTC 04:00
    from datetime import time as dtime
    import pytz
    app.job_queue.run_daily(
        otomatik_guncelle,
        time=dtime(hour=4, minute=0, tzinfo=pytz.utc),
    )

    print("✅ Bot çalışıyor! Telegram'da /start yaz.")
    app.run_polling()

if __name__ == "__main__":
    main()
