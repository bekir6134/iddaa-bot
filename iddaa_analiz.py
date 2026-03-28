import requests
import openpyxl
import time
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

FD_KEY   = "0fb7930eaa3744f4b6175315a4b734ce"
ODDS_KEY = "ae7313d30cccf1080f66e93c57d9da7a"

FD_BASE   = "https://api.football-data.org/v4"
ODDS_BASE = "https://api.the-odds-api.com/v4"

LIGLER = {
    "Premier League": {"fd_code": "PL",  "odds_key": "soccer_epl"},
    "La Liga":        {"fd_code": "PD",  "odds_key": "soccer_spain_la_liga"},
    "Bundesliga":     {"fd_code": "BL1", "odds_key": "soccer_germany_bundesliga"},
    "Serie A":        {"fd_code": "SA",  "odds_key": "soccer_italy_serie_a"},
    "Ligue 1":        {"fd_code": "FL1", "odds_key": "soccer_france_ligue_one"},
    "Süper Lig":      {"fd_code": "TSL", "odds_key": "soccer_turkey_super_league"},
}

RENKLER = {
    "baslik_bg":   "1F4E79",
    "baslik_yazi": "FFFFFF",
    "alt_baslik":  "2E75B6",
    "satir1":      "DEEAF1",
    "satir2":      "FFFFFF",
    "kazandi":     "C6EFCE",
    "kaybetti":    "FFC7CE",
    "beraberlik":  "FFEB9C",
    "pozitif":     "375623",
    "negatif":     "9C0006",
}

_fd_call_count = 0

def fd_get(endpoint, params=None):
    global _fd_call_count
    # Football-Data.org free tier: 10 req/dakika limiti — her çağrıda 7sn bekle
    if _fd_call_count > 0:
        time.sleep(7)
    _fd_call_count += 1
    r = requests.get(f"{FD_BASE}/{endpoint}",
                     headers={"X-Auth-Token": FD_KEY},
                     params=params, timeout=20)
    if r.status_code == 200:
        return r.json()
    print(f"  FD API hata {r.status_code}: {endpoint}")
    return None

def odds_get(sport_key):
    """Tek çağrıda h2h + totals + btts + h2h_h1 + asian_handicap marketlerini çek."""
    r = requests.get(f"{ODDS_BASE}/sports/{sport_key}/odds",
                     params={
                         "apiKey": ODDS_KEY,
                         "regions": "eu",
                         "markets": "h2h,totals,btts,h2h_h1,asian_handicap",
                         "oddsFormat": "decimal",
                     },
                     timeout=20)
    if r.status_code == 200:
        return r.json()
    print(f"  Odds API hata {r.status_code}: {sport_key}")
    return []

# ── Yardımcı: Market ortalama hesapla ────────────────────────────────────────
def _ort(bookmakers, market_key, outcome_name):
    vals = []
    for bm in bookmakers:
        for mkt in bm.get("markets", []):
            if mkt.get("key") == market_key:
                for o in mkt.get("outcomes", []):
                    if o.get("name", "").lower() == outcome_name.lower():
                        vals.append(o["price"])
    return round(sum(vals) / len(vals), 2) if vals else None

def _ort_totals(bookmakers, point, over_under):
    """Totals marketi için belirli puan (2.5 gibi) ve Over/Under ortalaması."""
    vals = []
    for bm in bookmakers:
        for mkt in bm.get("markets", []):
            if mkt.get("key") == "totals":
                for o in mkt.get("outcomes", []):
                    if (o.get("name", "").lower() == over_under.lower() and
                            abs(float(o.get("point", 0)) - point) < 0.01):
                        vals.append(o["price"])
    return round(sum(vals) / len(vals), 2) if vals else None

def _ort_handicap(bookmakers, home_team):
    """Asian handicap: ev sahibi ve deplasman oranı + handicap değeri."""
    ev_vals = []; dep_vals = []; hcp_vals = []
    for bm in bookmakers:
        for mkt in bm.get("markets", []):
            if mkt.get("key") == "asian_handicap":
                for o in mkt.get("outcomes", []):
                    if o.get("name", "") == home_team:
                        ev_vals.append(o["price"])
                        hcp_vals.append(float(o.get("point", 0)))
                    else:
                        dep_vals.append(o["price"])
    ev  = round(sum(ev_vals)  / len(ev_vals),  2) if ev_vals  else None
    dep = round(sum(dep_vals) / len(dep_vals), 2) if dep_vals else None
    hcp = round(sum(hcp_vals) / len(hcp_vals), 2) if hcp_vals else None
    return ev, dep, hcp

# ── H2H: Tarihsel maç listesinden iki takımın karşılaşmalarını bul ────────────
def h2h_hesapla(maclar, ev_takim, dep_takim, son_n=5):
    """
    maclar: fd_get'ten gelen 'matches' listesi (tamamlananlar)
    Döndürür: son_n karşılaşmayı içeren liste [{tarih, ev, dep, skor, sonuc}]
    """
    bulunanlar = []
    for m in maclar:
        ht = m.get("homeTeam", {}).get("name", "")
        at = m.get("awayTeam", {}).get("name", "")
        if not ((ev_takim.lower() in ht.lower() or ev_takim.lower() in at.lower()) and
                (dep_takim.lower() in ht.lower() or dep_takim.lower() in at.lower())):
            continue
        full = m.get("score", {}).get("fullTime", {})
        hg = full.get("home"); ag = full.get("away")
        if hg is None:
            continue
        if hg > ag:   sonuc = "1"
        elif hg == ag: sonuc = "X"
        else:          sonuc = "2"
        bulunanlar.append({
            "tarih": m.get("utcDate", "")[:10],
            "ev": ht, "dep": at,
            "skor": f"{hg}-{ag}",
            "sonuc": sonuc,
            "toplam": hg + ag,
        })
    return bulunanlar[-son_n:]

# ── Takım form: son N maçta ev/dep ayrımıyla performans ──────────────────────
def takim_form_hesapla(maclar, takim, son_n=5):
    """
    Döndürür: {
      "form": "G-G-B-M-G",
      "galibiyet": int, "beraberlik": int, "maglubiyet": int,
      "avg_att": float, "avg_yedi": float,
      "ev_form": str, "dep_form": str
    }
    """
    ilgili = []
    for m in maclar:
        ht = m.get("homeTeam", {}).get("name", "")
        at = m.get("awayTeam", {}).get("name", "")
        if takim.lower() not in ht.lower() and takim.lower() not in at.lower():
            continue
        full = m.get("score", {}).get("fullTime", {})
        hg = full.get("home"); ag = full.get("away")
        if hg is None:
            continue
        ev_mi = takim.lower() in ht.lower()
        att  = hg if ev_mi else ag
        yedi = ag if ev_mi else hg
        if att > yedi:   sonuc = "G"
        elif att == yedi: sonuc = "B"
        else:             sonuc = "M"
        ilgili.append({"sonuc": sonuc, "att": att, "yedi": yedi, "ev_mi": ev_mi})

    son = ilgili[-son_n:]
    if not son:
        return {"form": "-", "galibiyet": 0, "beraberlik": 0, "maglubiyet": 0,
                "avg_att": 0.0, "avg_yedi": 0.0, "ev_form": "-", "dep_form": "-"}

    g = sum(1 for x in son if x["sonuc"] == "G")
    b = sum(1 for x in son if x["sonuc"] == "B")
    m_ = sum(1 for x in son if x["sonuc"] == "M")
    avg_att  = round(sum(x["att"]  for x in son) / len(son), 1)
    avg_yedi = round(sum(x["yedi"] for x in son) / len(son), 1)
    form_str = "-".join(x["sonuc"] for x in son)

    ev_maclar  = [x for x in son if x["ev_mi"]]
    dep_maclar = [x for x in son if not x["ev_mi"]]
    ev_form  = "-".join(x["sonuc"] for x in ev_maclar)  or "-"
    dep_form = "-".join(x["sonuc"] for x in dep_maclar) or "-"

    return {
        "form": form_str, "galibiyet": g, "beraberlik": b, "maglubiyet": m_,
        "avg_att": avg_att, "avg_yedi": avg_yedi,
        "ev_form": ev_form, "dep_form": dep_form,
    }

# ── Stil yardımcıları ─────────────────────────────────────────────────────────
def stil_baslik(ws, row, col, deger, bg=None, yazi=None, bold=True, boyut=11, hizala="center"):
    c = ws.cell(row=row, column=col, value=deger)
    c.font = Font(name="Arial", bold=bold, color=yazi or RENKLER["baslik_yazi"], size=boyut)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=hizala, vertical="center", wrap_text=True)
    return c

def stil_veri(ws, row, col, deger, bg=None, yazi="000000", bold=False, hizala="center", sayi_fmt=None):
    c = ws.cell(row=row, column=col, value=deger)
    c.font = Font(name="Arial", bold=bold, color=yazi, size=10)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=hizala, vertical="center")
    if sayi_fmt:
        c.number_format = sayi_fmt
    return c

def ince_kenar(ws, min_row, max_row, min_col, max_col):
    ince = Side(style="thin", color="BFBFBF")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for c in row:
            c.border = Border(left=ince, right=ince, top=ince, bottom=ince)

def col_gen(ws, sutunlar):
    for i, en in enumerate(sutunlar, 1):
        ws.column_dimensions[get_column_letter(i)].width = en

# ── Sayfa: Puan Tablosu ───────────────────────────────────────────────────────
def yaz_puan_tablosu(ws, lig_adi, fd_code, mac_verisi=None):
    print(f"  Puan tablosu çekiliyor: {lig_adi}")
    data = fd_get(f"competitions/{fd_code}/standings")
    if not data:
        ws["A1"] = "Veri alınamadı"
        return

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A1:M1")
    stil_baslik(ws, 1, 1, f"{lig_adi} — Puan Tablosu ({datetime.now().strftime('%d.%m.%Y')})",
                bg=RENKLER["baslik_bg"], boyut=13)

    basliklar = ["Sıra", "Takım", "O", "G", "B", "M", "AG", "YG", "A", "Puan",
                 "Form", "Ev G%", "Dep G%"]
    for i, b in enumerate(basliklar, 1):
        stil_baslik(ws, 2, i, b, bg=RENKLER["alt_baslik"], boyut=10)

    col_gen(ws, [6, 22, 5, 5, 5, 5, 6, 6, 6, 7, 12, 8, 8])

    standings = data.get("standings", [{}])[0].get("table", [])

    # Maç verisi varsa ev/dep galibiyetlerini hesapla
    ev_dep_stats = {}
    if mac_verisi:
        for m in mac_verisi:
            full = m.get("score", {}).get("fullTime", {})
            hg = full.get("home"); ag = full.get("away")
            if hg is None:
                continue
            ht = m.get("homeTeam", {}).get("name", "")
            at = m.get("awayTeam", {}).get("name", "")
            for t in [ht, at]:
                if t not in ev_dep_stats:
                    ev_dep_stats[t] = {"ev_g": 0, "ev_mac": 0, "dep_g": 0, "dep_mac": 0}
            ev_dep_stats[ht]["ev_mac"] += 1
            ev_dep_stats[at]["dep_mac"] += 1
            if hg > ag:
                ev_dep_stats[ht]["ev_g"] += 1
            elif ag > hg:
                ev_dep_stats[at]["dep_g"] += 1

    for idx, t in enumerate(standings):
        row = idx + 3
        bg = RENKLER["satir1"] if idx % 2 == 0 else RENKLER["satir2"]
        pos = t.get("position", "")

        if pos <= 1:   bg_pos = "FFD700"
        elif pos <= 4: bg_pos = "C6EFCE"
        elif pos >= len(standings) - 2: bg_pos = "FFC7CE"
        else:          bg_pos = bg

        takim_adi = t.get("team", {}).get("name", "")
        stil_veri(ws, row, 1,  pos,                              bg=bg_pos, bold=True)
        stil_veri(ws, row, 2,  takim_adi,                        bg=bg, hizala="left", bold=True)
        stil_veri(ws, row, 3,  t.get("playedGames", ""),         bg=bg)
        stil_veri(ws, row, 4,  t.get("won", ""),                 bg=bg)
        stil_veri(ws, row, 5,  t.get("draw", ""),                bg=bg)
        stil_veri(ws, row, 6,  t.get("lost", ""),                bg=bg)
        stil_veri(ws, row, 7,  t.get("goalsFor", ""),            bg=bg)
        stil_veri(ws, row, 8,  t.get("goalsAgainst", ""),        bg=bg)
        stil_veri(ws, row, 9,  t.get("goalDifference", ""),      bg=bg)
        stil_veri(ws, row, 10, t.get("points", ""),              bg=bg, bold=True)
        stil_veri(ws, row, 11, t.get("form", "") or "-",         bg=bg)

        # Ev / Deplasman galibiyet oranı — maç verisinden hesapla
        stats = ev_dep_stats.get(takim_adi, {})
        ev_mac  = stats.get("ev_mac", 0) or 1
        dep_mac = stats.get("dep_mac", 0) or 1
        ev_pct  = round(stats.get("ev_g", 0) / ev_mac * 100, 1) if stats else None
        dep_pct = round(stats.get("dep_g", 0) / dep_mac * 100, 1) if stats else None

        if ev_pct is not None:
            stil_veri(ws, row, 12, f"%{ev_pct}", bg=bg)
            stil_veri(ws, row, 13, f"%{dep_pct}", bg=bg)
        else:
            for c in [12, 13]:
                ws.cell(row=row, column=c, value="-").alignment = Alignment(horizontal="center")

    ince_kenar(ws, 2, len(standings) + 2, 1, 13)

# ── Sayfa: Maç Sonuçları ──────────────────────────────────────────────────────
def yaz_mac_sonuclari(ws, lig_adi, maclar):
    print(f"  Maç sonuçları yazılıyor: {lig_adi}")

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A1:J1")
    stil_baslik(ws, 1, 1, f"{lig_adi} — Son 50 Maç Sonucu",
                bg=RENKLER["baslik_bg"], boyut=13)

    basliklar = ["Tarih", "Hafta", "Ev Sahibi", "Skor", "Deplasman",
                 "Sonuç", "İY Skor", "Toplam Gol", "KG", "Durum"]
    for i, b in enumerate(basliklar, 1):
        stil_baslik(ws, 2, i, b, bg=RENKLER["alt_baslik"], boyut=10)

    col_gen(ws, [12, 7, 22, 8, 22, 8, 9, 11, 6, 10])

    for idx, m in enumerate(maclar):
        row = idx + 3
        bg = RENKLER["satir1"] if idx % 2 == 0 else RENKLER["satir2"]

        tarih = m.get("utcDate", "")[:10]
        hafta = m.get("matchday", "")
        ev    = m.get("homeTeam", {}).get("name", "")
        dep   = m.get("awayTeam", {}).get("name", "")
        full  = m.get("score", {}).get("fullTime", {})
        ht    = m.get("score", {}).get("halfTime", {})
        ev_g  = full.get("home"); dep_g = full.get("away")
        iy_ev = ht.get("home");   iy_dep = ht.get("away")

        if ev_g is not None and dep_g is not None:
            skor   = f"{ev_g} - {dep_g}"
            iy     = f"{iy_ev} - {iy_dep}" if iy_ev is not None else "-"
            toplam = ev_g + dep_g
            kg     = "Var" if ev_g > 0 and dep_g > 0 else "Yok"
            if ev_g > dep_g:    sonuc = "1"; bg_s = RENKLER["kazandi"]
            elif ev_g == dep_g: sonuc = "X"; bg_s = RENKLER["beraberlik"]
            else:               sonuc = "2"; bg_s = RENKLER["kaybetti"]
        else:
            skor = iy = "-"; toplam = ""; kg = ""; sonuc = "-"; bg_s = bg

        stil_veri(ws, row, 1,  tarih,  bg=bg, hizala="center")
        stil_veri(ws, row, 2,  hafta,  bg=bg)
        stil_veri(ws, row, 3,  ev,     bg=bg, hizala="left")
        stil_veri(ws, row, 4,  skor,   bg=bg, bold=True)
        stil_veri(ws, row, 5,  dep,    bg=bg, hizala="left")
        stil_veri(ws, row, 6,  sonuc,  bg=bg_s, bold=True)
        stil_veri(ws, row, 7,  iy,     bg=bg)
        stil_veri(ws, row, 8,  toplam, bg=bg)
        stil_veri(ws, row, 9,  kg,     bg=bg)
        stil_veri(ws, row, 10, m.get("status", ""), bg=bg)

    ince_kenar(ws, 2, len(maclar) + 2, 1, 10)

# ── Sayfa: Bahis Oranları (Genişletilmiş) ────────────────────────────────────
def yaz_oranlar(ws, lig_adi, odds_key):
    print(f"  Bahis oranları çekiliyor: {lig_adi}")
    maclar = odds_get(odds_key)

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # Sütun başlıkları — 20 sütun
    basliklar = [
        "Tarih", "Saat", "Ev Sahibi", "Deplasman",
        "1", "X", "2",
        "1%", "X%", "2%", "Marj%", "Favori",
        "Over2.5", "Under2.5", "KG Evet", "KG Hayır",
        "İY-1", "İY-X", "İY-2",
        "Handicap Ev", "Handicap Dep", "HCP Değer",
    ]
    toplam_sutun = len(basliklar)

    ws.merge_cells(f"A1:{get_column_letter(toplam_sutun)}1")
    stil_baslik(ws, 1, 1, f"{lig_adi} — Yaklaşan Maç Oranları",
                bg=RENKLER["baslik_bg"], boyut=13)

    for i, b in enumerate(basliklar, 1):
        stil_baslik(ws, 2, i, b, bg=RENKLER["alt_baslik"], boyut=10)

    col_gen(ws, [11, 7, 22, 22,
                 7, 7, 7, 8, 8, 8, 7, 14,
                 9, 9, 8, 8,
                 7, 7, 7,
                 10, 10, 8])

    if not maclar:
        ws.cell(row=3, column=1, value="Veri bulunamadı (yaklaşan maç yok veya API limiti)")
        return

    for idx, m in enumerate(maclar):
        row = idx + 3
        bg = RENKLER["satir1"] if idx % 2 == 0 else RENKLER["satir2"]

        dt    = m.get("commence_time", "")
        tarih = dt[:10] if dt else ""
        saat  = dt[11:16] if len(dt) > 15 else ""
        ev    = m.get("home_team", "")
        dep   = m.get("away_team", "")
        bms   = m.get("bookmakers", [])

        # 1X2
        o1 = _ort(bms, "h2h", ev)
        ox = _ort(bms, "h2h", "Draw")
        o2 = _ort(bms, "h2h", dep)

        # Totals 2.5
        over25  = _ort_totals(bms, 2.5, "Over")
        under25 = _ort_totals(bms, 2.5, "Under")

        # BTTS
        kg_evet  = _ort(bms, "btts", "Yes")
        kg_hayir = _ort(bms, "btts", "No")

        # İlk Yarı 1X2
        iy1 = _ort(bms, "h2h_h1", ev)
        iyx = _ort(bms, "h2h_h1", "Draw")
        iy2 = _ort(bms, "h2h_h1", dep)

        # Handicap
        hcp_ev, hcp_dep, hcp_val = _ort_handicap(bms, ev)

        stil_veri(ws, row, 1, tarih, bg=bg)
        stil_veri(ws, row, 2, saat,  bg=bg)
        stil_veri(ws, row, 3, ev,    bg=bg, hizala="left")
        stil_veri(ws, row, 4, dep,   bg=bg, hizala="left")

        if o1 and ox and o2:
            min_o = min(o1, ox, o2)
            stil_veri(ws, row, 5,  o1, bg="C6EFCE" if o1 == min_o else bg, bold=o1 == min_o, sayi_fmt="0.00")
            stil_veri(ws, row, 6,  ox, bg="C6EFCE" if ox == min_o else bg, bold=ox == min_o, sayi_fmt="0.00")
            stil_veri(ws, row, 7,  o2, bg="C6EFCE" if o2 == min_o else bg, bold=o2 == min_o, sayi_fmt="0.00")
            p1 = 1/o1; px = 1/ox; p2 = 1/o2; ptop = p1 + px + p2
            stil_veri(ws, row, 8,  round(p1/ptop*100, 1), bg=bg, sayi_fmt="0.0")
            stil_veri(ws, row, 9,  round(px/ptop*100, 1), bg=bg, sayi_fmt="0.0")
            stil_veri(ws, row, 10, round(p2/ptop*100, 1), bg=bg, sayi_fmt="0.0")
            marj = round((ptop - 1)*100, 1)
            stil_veri(ws, row, 11, marj, bg=bg, sayi_fmt="0.0")
            favori = ev if o1 == min_o else ("Beraberlik" if ox == min_o else dep)
            stil_veri(ws, row, 12, favori, bg=bg, bold=True)
        else:
            for c in range(5, 13):
                ws.cell(row=row, column=c, value="-").alignment = Alignment(horizontal="center")

        # Totals
        stil_veri(ws, row, 13, over25  or "-", bg=bg, sayi_fmt="0.00" if over25  else None)
        stil_veri(ws, row, 14, under25 or "-", bg=bg, sayi_fmt="0.00" if under25 else None)

        # BTTS
        stil_veri(ws, row, 15, kg_evet  or "-", bg=bg, sayi_fmt="0.00" if kg_evet  else None)
        stil_veri(ws, row, 16, kg_hayir or "-", bg=bg, sayi_fmt="0.00" if kg_hayir else None)

        # İlk Yarı
        stil_veri(ws, row, 17, iy1 or "-", bg=bg, sayi_fmt="0.00" if iy1 else None)
        stil_veri(ws, row, 18, iyx or "-", bg=bg, sayi_fmt="0.00" if iyx else None)
        stil_veri(ws, row, 19, iy2 or "-", bg=bg, sayi_fmt="0.00" if iy2 else None)

        # Handicap
        stil_veri(ws, row, 20, hcp_ev  or "-", bg=bg, sayi_fmt="0.00" if hcp_ev  else None)
        stil_veri(ws, row, 21, hcp_dep or "-", bg=bg, sayi_fmt="0.00" if hcp_dep else None)
        stil_veri(ws, row, 22, f"{hcp_val:+.1f}" if hcp_val is not None else "-", bg=bg)

    ince_kenar(ws, 2, len(maclar) + 2, 1, toplam_sutun)

# ── Sayfa: İstatistik Özet (H2H + Form eklenmiş) ─────────────────────────────
def yaz_istatistik(ws, lig_adi, maclar):
    print(f"  İstatistikler hesaplanıyor: {lig_adi}")

    tamam = [m for m in maclar
             if m.get("score", {}).get("fullTime", {}).get("home") is not None]

    if not tamam:
        ws["A1"] = "Tamamlanan maç yok"; return

    n = len(tamam)
    ev_g = dep_g = beraberlik = toplam_gol = kg_var = iki_yas = uc_yas = 0
    for m in tamam:
        h = m["score"]["fullTime"]["home"]
        a = m["score"]["fullTime"]["away"]
        if h > a:   ev_g += 1
        elif h == a: beraberlik += 1
        else:        dep_g += 1
        toplam_gol += h + a
        if h > 0 and a > 0: kg_var += 1
        if h + a > 2: iki_yas += 1
        if h + a > 3: uc_yas += 1

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A1:D1")
    stil_baslik(ws, 1, 1, f"{lig_adi} — Sezon İstatistikleri ({n} maç)",
                bg=RENKLER["baslik_bg"], boyut=13)

    for i, b in enumerate(["İstatistik", "Değer", "Yüzde", "Yorum"], 1):
        stil_baslik(ws, 2, i, b, bg=RENKLER["alt_baslik"], boyut=10)

    col_gen(ws, [28, 10, 10, 25])

    satirlar = [
        ("Toplam Maç",            n,             None,            ""),
        ("Ev Sahibi Galibiyeti",  ev_g,          ev_g/n,          "1 oynamak için baz oran"),
        ("Beraberlik",            beraberlik,    beraberlik/n,    "X için baz oran"),
        ("Deplasman Galibiyeti",  dep_g,         dep_g/n,         "2 oynamak için baz oran"),
        ("Maç Başı Ort. Gol",     round(toplam_gol/n, 2), None,  ""),
        ("2.5 Üst (3+ gol)",      iki_yas,       iki_yas/n,       "2.5 üst bahisi için"),
        ("2.5 Alt (0-2 gol)",     n - iki_yas,   (n - iki_yas)/n, "2.5 alt bahisi için"),
        ("3.5 Üst (4+ gol)",      uc_yas,        uc_yas/n,        ""),
        ("Karşılıklı Gol (KG)",   kg_var,        kg_var/n,        "Her iki takım da gol attı"),
        ("KG Yok",                n - kg_var,    (n - kg_var)/n,  ""),
    ]

    for idx, (isim, deger, pct, yorum) in enumerate(satirlar):
        row = idx + 3
        bg = RENKLER["satir1"] if idx % 2 == 0 else RENKLER["satir2"]
        stil_veri(ws, row, 1, isim,  bg=bg, hizala="left", bold=True)
        stil_veri(ws, row, 2, deger, bg=bg)
        if pct is not None:
            stil_veri(ws, row, 3, pct, bg=bg, sayi_fmt="0.0%")
        else:
            ws.cell(row=row, column=3, value="-").alignment = Alignment(horizontal="center")
        stil_veri(ws, row, 4, yorum, bg=bg, hizala="left")

    ince_kenar(ws, 2, len(satirlar) + 2, 1, 4)

    # ── Takım bazlı gol istatistikleri ────────────────────────────────────────
    takim_stats = {}
    for m in tamam:
        ev_t  = m.get("homeTeam", {}).get("name", "")
        dep_t = m.get("awayTeam", {}).get("name", "")
        h = m["score"]["fullTime"]["home"]
        a = m["score"]["fullTime"]["away"]
        for t, gol_at, gol_ye, ev_mi in [(ev_t, h, a, True), (dep_t, a, h, False)]:
            if t not in takim_stats:
                takim_stats[t] = {"gol_at": 0, "gol_ye": 0, "mac": 0,
                                  "ev_mac": 0, "ev_g": 0, "dep_mac": 0, "dep_g": 0}
            takim_stats[t]["gol_at"] += gol_at
            takim_stats[t]["gol_ye"] += gol_ye
            takim_stats[t]["mac"] += 1
            if ev_mi:
                takim_stats[t]["ev_mac"] += 1
                if h > a: takim_stats[t]["ev_g"] += 1
            else:
                takim_stats[t]["dep_mac"] += 1
                if a > h: takim_stats[t]["dep_g"] += 1

    bslk_row = len(satirlar) + 5
    ws.merge_cells(f"A{bslk_row}:H{bslk_row}")
    stil_baslik(ws, bslk_row, 1, "Takım Gol & Performans İstatistikleri",
                bg=RENKLER["alt_baslik"], boyut=11)

    takim_basliklar = ["Takım", "Maç", "Gol Attı", "Gol Yedi", "Avg Att", "Avg Yedi", "Ev G%", "Dep G%"]
    for i, b in enumerate(takim_basliklar, 1):
        stil_baslik(ws, bslk_row + 1, i, b, bg=RENKLER["alt_baslik"], boyut=10)

    sirali = sorted(takim_stats.items(), key=lambda x: -x[1]["gol_at"])
    for idx, (takim, s) in enumerate(sirali):
        row = bslk_row + 2 + idx
        bg = RENKLER["satir1"] if idx % 2 == 0 else RENKLER["satir2"]
        mac = s["mac"] or 1
        ev_pct  = round(s["ev_g"]  / (s["ev_mac"]  or 1) * 100, 1)
        dep_pct = round(s["dep_g"] / (s["dep_mac"] or 1) * 100, 1)
        stil_veri(ws, row, 1, takim,                             bg=bg, hizala="left")
        stil_veri(ws, row, 2, s["mac"],                          bg=bg)
        stil_veri(ws, row, 3, s["gol_at"],                       bg=bg)
        stil_veri(ws, row, 4, s["gol_ye"],                       bg=bg)
        stil_veri(ws, row, 5, round(s["gol_at"] / mac, 2),       bg=bg, sayi_fmt="0.00")
        stil_veri(ws, row, 6, round(s["gol_ye"] / mac, 2),       bg=bg, sayi_fmt="0.00")
        stil_veri(ws, row, 7, f"%{ev_pct}",                      bg=bg)
        stil_veri(ws, row, 8, f"%{dep_pct}",                     bg=bg)

    ince_kenar(ws, bslk_row, bslk_row + 1 + len(sirali), 1, 8)

    # ── Takım Form Tablosu (Son 5 maç) ───────────────────────────────────────
    form_row = bslk_row + 2 + len(sirali) + 3
    ws.merge_cells(f"A{form_row}:H{form_row}")
    stil_baslik(ws, form_row, 1, "Takım Form Analizi (Son 5 Maç)",
                bg=RENKLER["alt_baslik"], boyut=11)

    form_basliklar = ["Takım", "Son 5 Form", "G", "B", "M", "Avg Att", "Avg Yedi", "Ev Form / Dep Form"]
    for i, b in enumerate(form_basliklar, 1):
        stil_baslik(ws, form_row + 1, i, b, bg=RENKLER["alt_baslik"], boyut=10)

    takimlar_sirali = sorted(takim_stats.keys())
    for idx, takim in enumerate(takimlar_sirali):
        row = form_row + 2 + idx
        bg = RENKLER["satir1"] if idx % 2 == 0 else RENKLER["satir2"]
        f = takim_form_hesapla(tamam, takim, son_n=5)
        stil_veri(ws, row, 1, takim,             bg=bg, hizala="left")
        stil_veri(ws, row, 2, f["form"],          bg=bg)
        stil_veri(ws, row, 3, f["galibiyet"],     bg=bg)
        stil_veri(ws, row, 4, f["beraberlik"],    bg=bg)
        stil_veri(ws, row, 5, f["maglubiyet"],    bg=bg)
        stil_veri(ws, row, 6, f["avg_att"],       bg=bg, sayi_fmt="0.0")
        stil_veri(ws, row, 7, f["avg_yedi"],      bg=bg, sayi_fmt="0.0")
        stil_veri(ws, row, 8, f"{f['ev_form']} / {f['dep_form']}", bg=bg, hizala="left")

    ince_kenar(ws, form_row, form_row + 1 + len(takimlar_sirali), 1, 8)

# ── Özet Sayfası ─────────────────────────────────────────────────────────────
def yaz_ozet(ws):
    ws.merge_cells("A1:F1")
    stil_baslik(ws, 1, 1,
                f"İddaa Analiz Sistemi — Güncelleme: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
                bg=RENKLER["baslik_bg"], boyut=14)
    ws.row_dimensions[1].height = 35

    for i, b in enumerate(["Lig", "Sekme", "İçerik", "Son Güncelleme"], 1):
        stil_baslik(ws, 2, i, b, bg=RENKLER["alt_baslik"])

    col_gen(ws, [20, 25, 55, 20])

    lig_listesi = list(LIGLER.keys())
    kisaltmalar = {"Premier League": "PL", "La Liga": "LL", "Bundesliga": "BL",
                   "Serie A": "SA", "Ligue 1": "L1", "Süper Lig": "SL"}

    sekmeler = []
    for lig in lig_listesi:
        kisa = kisaltmalar.get(lig, lig[:2])
        sekmeler += [
            (lig, f"{kisa} - Puan Tablosu",  "Sıralama, form, gol farkı, ev/dep %"),
            (lig, f"{kisa} - Maçlar",        "Son 50 maç, skor, 1X2, İY, KG"),
            (lig, f"{kisa} - Oranlar",       "1X2, Alt/Üst, KG, İY 1X2, Handikap oranları"),
            (lig, f"{kisa} - İstatistik",    "1X2%, 2.5 üst/alt, KG, takım form analizi"),
        ]

    guncelleme = datetime.now().strftime("%d.%m.%Y %H:%M")
    for idx, (lig, sekme, icerik) in enumerate(sekmeler):
        row = idx + 3
        bg = RENKLER["satir1"] if idx % 2 == 0 else RENKLER["satir2"]
        stil_veri(ws, row, 1, lig,         bg=bg, hizala="left", bold=True)
        stil_veri(ws, row, 2, sekme,       bg=bg, hizala="left")
        stil_veri(ws, row, 3, icerik,      bg=bg, hizala="left")
        stil_veri(ws, row, 4, guncelleme,  bg=bg)

    ince_kenar(ws, 2, len(sekmeler) + 2, 1, 4)

    not_row = len(sekmeler) + 5
    ws.merge_cells(f"A{not_row}:F{not_row}")
    stil_baslik(ws, not_row, 1, "NOTLAR", bg=RENKLER["alt_baslik"], boyut=11)
    notlar = [
        "• Bahis oranları birden fazla bookmaker'ın ortalamasıdır",
        "• İhtimal% = vig (marj) temizlenmiş gerçek olasılık tahmini",
        "• Marj% = bahisçinin karı (düşük marj = daha adil oran)",
        "• Alt/Üst ve KG oranları gerçek bahis marketinden çekilmiştir",
        "• H2H ve form hesaplamaları sezon içi maç verisine dayanır",
        "• İstatistikler mevcut sezon maçlarına dayanmaktadır",
    ]
    for i, n in enumerate(notlar):
        c = ws.cell(row=not_row + 1 + i, column=1, value=n)
        c.font = Font(name="Arial", size=10, italic=True)
        ws.merge_cells(f"A{not_row+1+i}:F{not_row+1+i}")

# ── ANA FONKSİYON ─────────────────────────────────────────────────────────────
def main():
    print("=" * 58)
    print("  İddaa Analiz Sistemi — Excel Raporu Oluşturuluyor")
    print("=" * 58)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_ozet = wb.create_sheet("OZET")
    yaz_ozet(ws_ozet)

    kisaltmalar = {
        "Premier League": "PL", "La Liga": "LL", "Bundesliga": "BL",
        "Serie A": "SA", "Ligue 1": "L1", "Süper Lig": "SL",
    }

    for lig_adi, bilgi in LIGLER.items():
        kisa = kisaltmalar.get(lig_adi, lig_adi[:2])
        print(f"\n[{lig_adi}]")

        # Maç verisini bir kez çek, hem Maçlar hem İstatistik hem Puan Tablosu için kullan
        print(f"  Maç sonuçları çekiliyor: {lig_adi}")
        mac_data = fd_get(f"competitions/{bilgi['fd_code']}/matches",
                          params={"status": "FINISHED", "limit": 100})
        mac_listesi = []
        if mac_data:
            mac_listesi = [m for m in mac_data.get("matches", [])
                           if m.get("score", {}).get("fullTime", {}).get("home") is not None]

        ws1 = wb.create_sheet(f"{kisa} - Puan Tablosu")
        yaz_puan_tablosu(ws1, lig_adi, bilgi["fd_code"], mac_verisi=mac_listesi)

        ws2 = wb.create_sheet(f"{kisa} - Maclar")
        yaz_mac_sonuclari(ws2, lig_adi, mac_listesi[:50])

        ws3 = wb.create_sheet(f"{kisa} - Oranlar")
        yaz_oranlar(ws3, lig_adi, bilgi["odds_key"])

        ws4 = wb.create_sheet(f"{kisa} - Istatistik")
        yaz_istatistik(ws4, lig_adi, mac_listesi)

    path = os.path.join(SCRIPT_DIR, "iddaa_analiz.xlsx")
    wb.save(path)
    print(f"\n{'='*58}")
    print(f"  Excel kaydedildi: {path}")
    print(f"  Toplam sayfa: {len(wb.sheetnames)}")
    print(f"{'='*58}")

if __name__ == "__main__":
    main()
